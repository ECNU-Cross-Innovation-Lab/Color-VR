import pandas as pd
from sklearn import preprocessing
from sklearn.model_selection import train_test_split

from interfaces import dataConstruct,subject
from pathlib import Path
import openpyxl
import re


raw_data=Path('A:/ZQJ/CILlab/VR/眼动和预测数据分析/data/VR教室数据终极版/VR_1')
pre_emotion=Path('A:\ZQJ\CILlab\VR\眼动和预测数据分析/data\论文资料整理\问卷\处理后的前测问卷.xlsx')
aft_emotion=Path('A:\ZQJ\CILlab\VR\眼动和预测数据分析/data\论文资料整理\问卷\处理后的后测问卷.xlsx')
reaction=Path('A:\ZQJ\CILlab\VR\眼动和预测数据分析/data\论文资料整理\正确率和反应时\反应时\反应时.xlsx')
states=['PR','DG','NG','NO','DO','DP','DY','NP','NY']
emotions=["prea_score","pred_score","afta_score","aftd_score"]
state_dict={'0':'PR','2':'DG','4':'NG','6':'NO','8':'DO','12':'DP','14':'DY','16':'NP','18':'NY'}
shuff_num = 50

label_mapping = { #映射表
    'mean_Tongue_Down':'舌头向下运动（均值）',
    "mean_Jaw_Right": "下颌向右运动（均值）",
    "mean_Jaw_Left": "下颌向左运动（均值）",
    "mean_Jaw_Forward": "下颌向前运动（均值）",
    "mean_Jaw_Open": "下颌张开（均值）",
    "mean_Mouth_Ape_Shape": "嘴巴形状（均值）",
    "mean_Mouth_Upper_Right": "上嘴唇右侧运动（均值）",
    "mean_Mouth_Upper_Left": "上嘴唇左侧运动（均值）",
    "mean_Mouth_Lower_Right": "下嘴唇右侧运动（均值）",
    "mean_Mouth_Lower_Left": "下嘴唇左侧运动（均值）",
    "mean_Mouth_Upper_Overturn": "上嘴唇翻转（均值）",
    "mean_Mouth_Lower_Overturn": "下嘴唇翻转（均值）",
    "mean_Mouth_Pout": "嘴巴撅嘴（均值）",
    "mean_Mouth_Smile_Right": "笑容右侧（均值）",
    "mean_Mouth_Smile_Left": "笑容左侧（均值）",
    "mean_Mouth_Sad_Right": "嘴唇右侧向下（均值）",
    "mean_Mouth_Sad_Left": "嘴唇左侧向下（均值）",
    "mean_Cheek_Puff_Right": "右脸颊鼓起（均值）",
    "mean_Cheek_Puff_Left": "左脸颊鼓起（均值）",
    "mean_Cheek_Suck": "脸颊吸入（均值）",
    "mean_Mouth_Upper_UpRight": "上嘴唇向右上方运动（均值）",
    "mean_Mouth_Upper_UpLeft": "上嘴唇向左上方运动（均值）",
    "mean_Mouth_Lower_DownRight": "下嘴唇向右下方运动（均值）",
    "mean_Mouth_Lower_DownLeft": "下嘴唇向左下方运动（均值）",
    "mean_Mouth_Upper_Inside": "上嘴唇吸吮（均值）",
    "mean_Mouth_Lower_Inside": "下嘴唇吸吮（均值）",
    "mean_Mouth_Lower_Overlay": "下嘴唇包裹上嘴唇（均值）",
    "mean_Tongue_LongStep1": "舌头运动点1（均值）",
    "mean_Tongue_LongStep2": "舌头运动点2（均值）",
    "mean_Tongue_Down_time": "舌头向下运动时间（均值）",
    "mean_Eye_Left_Blink": "左眼眨眼（均值）",
    "mean_Eye_Left_Wide": "左眼大小（均值）",
    "mean_Eye_Left_Right": "左眼向右运动（均值）",
    "mean_Eye_Left_Left": "左眼向左运动（均值）",
    "mean_Eye_Left_Up": "左眼向上运动（均值）",
    "mean_Eye_Left_Down": "左眼向下运动（均值）",
    "mean_Eye_Right_Blink": "右眼眨眼（均值）",
    "mean_Eye_Right_Wide": "右眼大小（均值）",
    "mean_Eye_Right_Right": "右眼向右运动（均值）",
    "mean_Eye_Right_Left": "右眼向左运动（均值）",
    "mean_Eye_Right_Up": "右眼向上运动（均值）",
    "mean_Eye_Right_Down": "右眼向下运动（均值）",
    "mean_Eye_Frown": "眉毛皱起（均值）",
    "mean_eyeOpenLeft": "左眼睁开（均值）",
    "mean_eyeOpenRight": "右眼睁开（均值）",
    "mean_pupilDiameterLeft": "左眼瞳孔直径（均值）",
    "mean_pupilDiameterRight": "右眼瞳孔直径（均值）",
    "var_Jaw_Right": "下颌向右运动（方差）",
    "var_Jaw_Left": "下颌向左运动（方差）",
    "var_Jaw_Forward": "下颌向前运动（方差）",
    "var_Jaw_Open": "下颌张开（方差）",
    "var_Mouth_Ape_Shape": "嘴巴形状（方差）",
    "var_Mouth_Upper_Right": "上嘴唇右侧运动（方差）",
    "var_Mouth_Upper_Left": "上嘴唇左侧运动（方差）",
    "var_Mouth_Lower_Right": "下嘴唇右侧运动（方差）",
    "var_Mouth_Lower_Left": "下嘴唇左侧运动（方差）",
    "var_Mouth_Upper_Overturn": "上嘴唇翻转（方差）",
    "var_Mouth_Lower_Overturn": "下嘴唇翻转（方差）",
    "var_Mouth_Pout": "嘴巴撅嘴（方差）",
    "var_Mouth_Smile_Right": "笑容右侧（方差）",
    "var_Mouth_Smile_Left": "笑容左侧（方差）",
    "var_Mouth_Sad_Right": "哀伤右侧（方差）",
    "var_Mouth_Sad_Left": "哀伤左侧（方差）",
    "var_Cheek_Puff_Right": "右脸颊鼓起（方差）",
    "var_Cheek_Puff_Left": "左脸颊鼓起（方差）",
    "var_Cheek_Suck": "脸颊吸入（方差）",
    "var_Mouth_Upper_UpRight": "上嘴唇向右上方运动（方差）",
    "var_Mouth_Upper_UpLeft": "上嘴唇向左上方运动（方差）",
    "var_Mouth_Lower_DownRight": "下嘴唇向右下方运动（方差）",
    "var_Mouth_Lower_DownLeft": "下嘴唇向左下方运动（方差）",
    "var_Mouth_Upper_Inside": "上嘴唇内部的运动（方差）",
    "var_Mouth_Lower_Inside": "下嘴唇内部的运动（方差）",
    "var_Mouth_Lower_Overlay": "下嘴唇覆盖（方差）",
    "var_Tongue_LongStep1": "舌头运动1（方差）",
    "var_Tongue_LongStep2": "舌头运动2（方差）",
    "var_Tongue_Down_time": "舌头向下运动的时间（方差）",
    "var_Eye_Left_Blink": "左眼眨眼（方差）",
    "var_Eye_Left_Wide": "左眼大小（方差）",
    "var_Eye_Left_Right": "左眼向右运动（方差）",
    "var_Eye_Left_Left": "左眼向左运动（方差）",
    "var_Eye_Left_Up": "左眼向上运动（方差）",
    "var_Eye_Left_Down": "左眼向下运动（方差）",
    "var_Eye_Right_Blink": "右眼眨眼（方差）",
    "var_Eye_Right_Wide": "右眼大小（方差）",
    "var_Eye_Right_Right": "右眼向右运动（方差）",
    "var_Eye_Right_Left": "右眼向左运动（方差）",
    "var_Eye_Right_Up": "右眼向上运动（方差）",
    "var_Eye_Right_Down": "右眼向下运动（方差）",
    "var_Eye_Frown": "眉毛皱起（方差）",
    "var_eyeOpenLeft": "左眼睁开（方差）",
    "var_eyeOpenRight": "右眼睁开（方差）",
    "var_pupilDiameterLeft": "左眼瞳孔直径（方差）",
    "var_pupilDiameterRight": "右眼瞳孔直径（方差）"
}


# pandas基于列event的值计算新的一列，用于标记场景
def transform_BG(value):
   # 正则表达式模式，匹配字符串末尾的数字
    pattern = r'\d+$'
    num=re.search(pattern,value).group()
    return state_dict[num]
class data_construct(dataConstruct):
    def __init__(self):
        self.datasets=dict()
        self.subjects=dict() #存储每个被试的数据。被试编号为键，值为对应的subject对象
    def load_data(self):
        for data in raw_data.iterdir():
            if data.is_dir():
                # 读入脸部和眼部数据和image状态数据
                # print(data.name)
                self.subjects[data.name]=subject()
                self.subjects[data.name].lip=pd.read_csv(data/'lip.csv')
                self.subjects[data.name].eye=pd.read_csv(data/'eye.csv')
                self.subjects[data.name].image=pd.read_csv(data/'image.csv')
                self.duplicated_delete(data.name)
                self.lip_eye_fusion(data.name)
                #读入情绪得分和反应时数据
        pre_emo=openpyxl.load_workbook(pre_emotion,data_only=True)
        aft_emo=openpyxl.load_workbook(aft_emotion,data_only=True)
        rea=openpyxl.load_workbook(reaction,data_only=True)
        # print(rea.sheetnames)
        for name in states[1:]:
            # wb_e=emo.get_sheet_by_name(name)
            wb_r=rea.get_sheet_by_name(name)
            #读取反应时数据
            start_r=2
            r_name=0
            while r_name!=None:
                r_name=wb_r.cell(row=1,column=start_r).value
                r_val=wb_r.cell(row=26,column=start_r).value
                # print(r_val)
                # print(r_name)
                # print(r_name,"  ",r_val)
                if r_val!=None:
                    self.subjects[r_name].r_time[name]=r_val
                    # print(r_name,"  ",r_val)
                    start_r+=1
    #   读取情绪得分数据
        pre_name=0
        prea_score=0
        pred_score=0
        afta_score=0
        aftd_score=0
        start_r=2
        while pre_name!=None:
            # print(pre_name," ",prea_score," ",pred_score," ",afta_score," ",aftd_score)
            pre_name=pre_emo.get_sheet_by_name('Sheet1').cell(row=start_r,column=2).value
            prea_score=pre_emo.get_sheet_by_name('Sheet1').cell(row=start_r,column=26).value
            pred_score=pre_emo.get_sheet_by_name('Sheet1').cell(row=start_r,column=27).value
            afta_score=aft_emo.get_sheet_by_name('Sheet1').cell(row=start_r,column=50).value
            aftd_score=aft_emo.get_sheet_by_name('Sheet1').cell(row=start_r,column=51).value
            if pre_name!=None:
                self.subjects[pre_name].emo_score["prea_score"]=prea_score
                self.subjects[pre_name].emo_score["pred_score"]=pred_score
                self.subjects[pre_name].emo_score["afta_score"]=afta_score
                self.subjects[pre_name].emo_score["aftd_score"]=aftd_score
                start_r+=1

    def state_decide(self,name,time,state):
        #根据传入的时间，比对对应的image，返回对应的分组状态
        #分组状态：
        # 2-DG
        pass
    def state_add(self):
    #     对于每个被试的状态数据，根据时间戳，将对应的状态加入到对应的脸部和眼部数据中
        pass
    def feature_clean(self,sample_rate=5):
        pass
    def feature_extract(self):
        pass

    def BG_feature_extract(self,sub):
        # 基于background对每个被试筛选对应数据，比计算对应特征的均值和方差
        # 其中部分数据为全0，后续考虑一下删除
        # print()
        col = list(self.subjects[sub].l_e.columns)
        # col.remove('BackGround')
        col.remove('event')
        col.remove('time')
        df=pd.DataFrame(self.subjects[sub].l_e,columns=col)


        bg_grouped=df.groupby('BackGround')
        for name,group in bg_grouped:
            col2 = list(group.columns)
            col2.remove('BackGround')
            g = pd.DataFrame(group, columns=col2)


            df_shuffle = g.sample(frac=1).reset_index(drop=True)
            # print("长度为:",len(df_shuffle))
            chunk_size = len(df_shuffle) //shuff_num

            for i in range(shuff_num):
                start = i * chunk_size
                end = (i + 1) * chunk_size if i < shuff_num - 1 else len(df_shuffle)
                gro = df_shuffle.iloc[start:end]

                gro_mean = gro.mean().to_frame().T
                gro_var = gro.var().to_frame().T
                # 修改列名并合并mean和var
                mean_col_dic = {name: 'mean_' + name for name in col2}
                var_col_dic = {name: 'var_' + name for name in col2}
                g_mean = gro_mean.rename(columns=mean_col_dic)
                g_var = gro_var.rename(columns=var_col_dic)

                if name in self.subjects[sub].r_time.keys():
                    r_time = pd.DataFrame({'r_time': self.subjects[sub].r_time[name]}, index=[0])
                    final_time = pd.concat([g_mean, g_var, r_time], axis=1)
                    # print(final_time)
                    # print(" ")
                    set_name_r = name + '_' + 'r_time'
                    # 标记对应的被试编号
                    sub_name=pd.DataFrame({'subject':sub},index=[0])
                    final_time=pd.concat([final_time,sub_name],axis=1)
                    self.datasets[set_name_r] = pd.concat(
                    [self.datasets[set_name_r], final_time], axis=0, ignore_index=True)
                for emo in emotions:
                        set_name = name+'_'+ emo
                        # print("name",sub)
                        # print(gro_mean)
                        sub_name = pd.DataFrame({'subject': sub}, index=[0])
                        pre=pd.DataFrame({emo:self.subjects[sub].emo_score[emo]},index=[0])
                        final_gro=pd.concat([g_mean,g_var,pre,sub_name],axis=1)
                        self.datasets[set_name]=pd.concat(
                            [self.datasets[set_name],final_gro],axis=0,ignore_index=True)

    def ALL_feature_extract(self,sub):
        col = list(self.subjects[sub].l_e.columns)
        col.remove('BackGround')
        col.remove('event')
        col.remove('time')
        df=pd.DataFrame(self.subjects[sub].l_e,columns=col)
        mean_col_dic = {name: 'mean_' + name for name in col}
        var_col_dic = {name: 'var_' + name for name in col}



        df_shuffle=df.sample(frac=1).reset_index(drop=True)
        chunk_size = len(df_shuffle) //shuff_num
        for i in range(shuff_num):
            start = i * chunk_size
            end = (i + 1) * chunk_size if i < shuff_num - 1 else len(df_shuffle)
            gro = df_shuffle.iloc[start:end]

            gro_mean = gro.mean().to_frame().T
            gro_var = gro.var().to_frame().T
            g_mean = gro_mean.rename(columns=mean_col_dic)
            g_var = gro_var.rename(columns=var_col_dic)

            for emo in emotions:
                sub_name = pd.DataFrame({'subject': sub}, index=[0])
                pre = pd.DataFrame({emo: self.subjects[sub].emo_score[emo]}, index=[0])
                set_name = 'ALL'+ '_' + emo
                final_gro = pd.concat([g_mean, g_var, pre,sub_name], axis=1)
                if final_gro.isna().sum().sum() == 0:
                    self.datasets[set_name] = pd.concat(
                        [self.datasets[set_name], final_gro], axis=0, ignore_index=True)
    def data_divide(self):
        # 对被试数据基于场景进行划分，形成对于前后正负情绪状态以及反应时预测的数据集
        # 基于场景划分数据————情绪状态
        # 命名格式：场景_情绪状态
        columns = list(self.subjects['m001'].l_e.columns)
        columns.remove('event')
        columns.remove('time')
        columns.remove('BackGround')

        for emo in emotions:
            col = ['mean_' + name for name in columns] + ['var_' + name for name in columns] + [emo]
            col2 = ['mean_' + name for name in columns] + ['var_' + name for name in columns] + ['r_time']
            self.datasets['ALL' + '_' + emo] = pd.DataFrame(columns=col)
            for state in states:
                # 构建state和emo对应的数据集格式
                self.datasets[state+'_'+emo]=pd.DataFrame(columns=col)
                if state!='PR':
                    self.datasets[state+'_'+'r_time']=pd.DataFrame(columns=col2)

        for sub in self.subjects.keys():
            self.ALL_feature_extract(sub)
            self.BG_feature_extract(sub)
    #                 基于Background选择对应数据

    #对每个被试的lip_eye数据加上场景标签，用于数据集的分类
    def BG_labeled(self,name):
        self.subjects[name].l_e['BackGround']=self.subjects[name].l_e['event'].apply(transform_BG)
    def duplicated_delete(self,subject_id):
        #删除重复数据
        self.subjects[subject_id].lip.drop_duplicates(subset='time', keep='first', inplace=True)
        self.subjects[subject_id].eye.drop_duplicates(subset='time', keep='first', inplace=True)
        self.subjects[subject_id].image.drop_duplicates(subset='time', keep='first', inplace=True)

    def lip_eye_fusion(self,subject_id):
        self.subjects[subject_id].lip.time = self.subjects[subject_id].lip.time.round(decimals=3)
        self.subjects[subject_id].eye.time = self.subjects[subject_id].eye.time.round(decimals=3)
        self.subjects[subject_id].image.time = self.subjects[subject_id].image.time.round(decimals=3)

        merged_data = pd.merge(self.subjects[subject_id].lip, self.subjects[subject_id].eye, on='time',how='inner')
        # print(subject_id)
        merged_data = pd.merge(merged_data, self.subjects[subject_id].image,  how='inner',on='time')
        self.subjects[subject_id].l_e=merged_data
        self.BG_labeled(subject_id)

    def datasetDivide_BySubject(self,ratio=0.7,dataset_name='ALL_pred_score'):
        #基于被试划分数据集
        #dataset_name为数据集的名称
        #ratio为训练集和测试集的比例,也代表了被试名的比例
        #返回训练集和测试集
        data=self.datasets[dataset_name]
        sub_list=data['subject'].unique()
        train_num=int(len(sub_list)*ratio)
        train_list=sub_list[:train_num]
        test_list=sub_list[train_num:]
        train_data=pd.DataFrame(columns=data.columns.drop('subject'))
        test_data=pd.DataFrame(columns=data.columns.drop('subject'))
        for sub in sub_list:
            if sub in train_list:
                sub_data=data[data['subject']==sub]
                sub_data=sub_data.drop(columns='subject')
                train_data=pd.concat([train_data,sub_data],axis=0,ignore_index=True)
            else:
                sub_data = data[data['subject'] == sub]
                sub_data = sub_data.drop(columns='subject')
                test_data = pd.concat([test_data, sub_data], axis=0, ignore_index=True)
        train_data=train_data.rename(columns=label_mapping)
        test_data=test_data.rename(columns=label_mapping)
        return train_data,test_data

